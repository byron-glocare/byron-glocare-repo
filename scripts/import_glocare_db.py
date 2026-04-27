"""
GLOCARE_DB.xlsx (유학 도메인 — 구글 드라이브 시트 export) → Supabase SQL.

7개 시트 → 7개 테이블:
  universities, departments, study_centers, study_cases,
  study_contacts, study_channels, study_insurance_claims

사용법:
    python scripts/import_glocare_db.py <xlsx-path> <out-sql>
"""

from __future__ import annotations

import datetime
import sys
from pathlib import Path

import openpyxl


# =============================================================================
# 시트 → 테이블 매핑 (header row + 컬럼 변환 규칙)
# =============================================================================

# 각 시트의 실제 헤더 row
HEADER_ROWS = {
    "universities": 4,
    "departments": 4,
    "centers": 3,
    "cases": 3,
    "contacts": 3,
    "channels": 3,
    "insurance_claims": 1,
}

# 시트 → 대상 테이블명
SHEET_TO_TABLE = {
    "universities": "universities",
    "departments": "departments",
    "centers": "study_centers",
    "cases": "study_cases",
    "contacts": "study_contacts",
    "channels": "study_channels",
    "insurance_claims": "study_insurance_claims",
}


# =============================================================================
# 유틸
# =============================================================================

def sql_quote(v) -> str:
    if v is None:
        return "NULL"
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v != v:
            return "NULL"
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v)
    if isinstance(v, datetime.datetime):
        # timestamptz literal
        return f"'{v.strftime('%Y-%m-%d %H:%M:%S')}+00'"
    if isinstance(v, datetime.date):
        return f"'{v.strftime('%Y-%m-%d')}'"
    s = str(v)
    return "'" + s.replace("\\", "\\\\").replace("'", "''") + "'"


def yn_to_bool(v):
    """Y/N text → boolean. 빈 값/None → False."""
    if v is None:
        return False
    s = str(v).strip().upper()
    return s == "Y"


def to_int(v):
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        try:
            return int(v)
        except (ValueError, OverflowError):
            return None
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        try:
            return int(float(s))
        except ValueError:
            return None
    return None


def to_str(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    return s


# =============================================================================
# 시트별 row 변환
# =============================================================================

def convert_universities_row(row: dict) -> dict | None:
    if not row.get("id"):
        return None
    return {
        "id": to_int(row.get("id")),
        "active": yn_to_bool(row.get("active")),
        "name_ko": to_str(row.get("name_ko")) or "(이름 없음)",
        "name_vi": to_str(row.get("name_vi")),
        "region_ko": to_str(row.get("region_ko")),
        "region_vi": to_str(row.get("region_vi")),
        "logo_url": to_str(row.get("logo_url")),
        "photo_url": to_str(row.get("photo_url")),
        "website_url": to_str(row.get("website_url")),
        "desc_ko": to_str(row.get("desc_ko")),
        "desc_vi": to_str(row.get("desc_vi")),
        "class_days_ko": to_str(row.get("class_days_ko")),
        "class_days_vi": to_str(row.get("class_days_vi")),
        "transport_bus": yn_to_bool(row.get("transport_bus")),
        "transport_subway": yn_to_bool(row.get("transport_subway")),
        "transport_train": yn_to_bool(row.get("transport_train")),
        "transport_desc_ko": to_str(row.get("transport_desc_ko")),
        "transport_desc_vi": to_str(row.get("transport_desc_vi")),
        "dormitory": yn_to_bool(row.get("dormitory")),
        "dormitory_desc_ko": to_str(row.get("dormitory_desc_ko")),
        "dormitory_desc_vi": to_str(row.get("dormitory_desc_vi")),
        "strengths": to_str(row.get("strengths")),
        "tags_ko": to_str(row.get("tags_ko")),
        "tags_vi": to_str(row.get("tags_vi")),
        "categories": to_str(row.get("categories")),
        "emoji": to_str(row.get("emoji")),
    }


def convert_departments_row(row: dict) -> dict | None:
    if not row.get("id") or not row.get("university_id"):
        return None
    return {
        "id": to_int(row.get("id")),
        "university_id": to_int(row.get("university_id")),
        "active": yn_to_bool(row.get("active")),
        "icon": to_str(row.get("icon")),
        "name_ko": to_str(row.get("name_ko")) or "(이름 없음)",
        "name_vi": to_str(row.get("name_vi")),
        "category": to_str(row.get("category")),
        "degree_years": to_int(row.get("degree_years")),
        "tuition_ko": to_str(row.get("tuition_ko")),
        "tuition_vi": to_str(row.get("tuition_vi")),
        "scholarship_ko": to_str(row.get("scholarship_ko")),
        "scholarship_vi": to_str(row.get("scholarship_vi")),
        "dept_url": to_str(row.get("dept_url")),
        "badge": to_str(row.get("badge")),
        "case_ids": to_str(row.get("case_ids")),
        "course": to_str(row.get("course")),
        "sort_order": to_int(row.get("sort_order")) or 0,
    }


def convert_centers_row(row: dict) -> dict | None:
    if not row.get("id"):
        return None
    return {
        "id": to_int(row.get("id")),
        "active": yn_to_bool(row.get("active")),
        "flag": to_str(row.get("flag")),
        "name_ko": to_str(row.get("name_ko")),
        "name_vi": to_str(row.get("name_vi")) or "(이름 없음)",
        "city_ko": to_str(row.get("city_ko")),
        "city_vi": to_str(row.get("city_vi")),
        "address": to_str(row.get("address")),
        "phone": to_str(row.get("phone")),
        "email": to_str(row.get("email")),
        "desc_ko": to_str(row.get("desc_ko")),
        "desc_vi": to_str(row.get("desc_vi")),
        "students_ko": to_str(row.get("students_ko")),
        "students_vi": to_str(row.get("students_vi")),
        "years_ko": to_str(row.get("years_ko")),
        "years_vi": to_str(row.get("years_vi")),
    }


def convert_cases_row(row: dict) -> dict | None:
    if not row.get("id"):
        return None
    return {
        "id": to_int(row.get("id")),
        "active": yn_to_bool(row.get("active")),
        "tiktok_url": to_str(row.get("tiktok_url")),
        "tiktok_thumb": to_str(row.get("tiktok_thumb")),
        "hero": yn_to_bool(row.get("hero")),
        "category_ko": to_str(row.get("category_ko")),
        "category_vi": to_str(row.get("category_vi")),
        "title_ko": to_str(row.get("title_ko")),
        "title_vi": to_str(row.get("title_vi")),
        "desc_ko": to_str(row.get("desc_ko")),
        "desc_vi": to_str(row.get("desc_vi")),
    }


def convert_contacts_row(row: dict) -> dict | None:
    if not row.get("id"):
        return None
    status = to_str(row.get("status")) or "미확인"
    if status not in ("미확인", "연락완료", "등록완료"):
        status = "미확인"
    return {
        "id": to_int(row.get("id")),
        "submitted_at": row.get("submitted_at"),
        "name": to_str(row.get("name")),
        "phone": to_str(row.get("phone")),
        "email": to_str(row.get("email")),
        "age": to_int(row.get("age")),
        "dept": to_str(row.get("dept")),
        "center": to_str(row.get("center")),
        "recruiting": to_str(row.get("recruiting")),
        "message": to_str(row.get("message")),
        "status": status,
        "memo": to_str(row.get("memo")),
    }


def convert_channels_row(row: dict) -> dict | None:
    if not row.get("id"):
        return None
    type_val = to_str(row.get("type"))
    if type_val and type_val not in (
        "tiktok", "facebook", "instagram", "youtube", "website", "kakao", "zalo", "other"
    ):
        type_val = "other"
    return {
        "id": to_int(row.get("id")),
        "active": yn_to_bool(row.get("active")),
        "type": type_val,
        "icon": to_str(row.get("icon")),
        "name_ko": to_str(row.get("name_ko")),
        "name_vi": to_str(row.get("name_vi")),
        "desc_ko": to_str(row.get("desc_ko")),
        "desc_vi": to_str(row.get("desc_vi")),
        "handle": to_str(row.get("handle")),
        "url": to_str(row.get("url")),
        "sort_order": to_int(row.get("sort_order")) or 0,
        "memo": to_str(row.get("memo")),
    }


def convert_insurance_row(row: dict) -> dict | None:
    if not row.get("id"):
        return None
    status = to_str(row.get("status")) or "미확인"
    if status not in ("미확인", "연락완료", "등록완료"):
        status = "미확인"
    return {
        "id": to_int(row.get("id")),
        "submitted_at": row.get("submitted_at"),
        "name": to_str(row.get("name")),
        "alien_no": to_str(row.get("alien_no")),
        "zalo": to_str(row.get("zalo")),
        "marketing": to_str(row.get("marketing")),
        "status": status,
        "memo": to_str(row.get("memo")),
    }


CONVERTERS = {
    "universities": convert_universities_row,
    "departments": convert_departments_row,
    "centers": convert_centers_row,
    "cases": convert_cases_row,
    "contacts": convert_contacts_row,
    "channels": convert_channels_row,
    "insurance_claims": convert_insurance_row,
}


# =============================================================================
# 시트 읽기
# =============================================================================

def read_sheet(wb, sheet_name: str) -> list[dict]:
    ws = wb[sheet_name]
    header_row = HEADER_ROWS[sheet_name]
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]

    rows = []
    converter = CONVERTERS[sheet_name]
    for r in range(header_row + 1, ws.max_row + 1):
        if not ws.cell(r, 1).value:
            continue
        raw = {}
        for c, h in enumerate(headers, 1):
            if h:
                raw[h] = ws.cell(r, c).value
        converted = converter(raw)
        if converted:
            rows.append(converted)
    return rows


# =============================================================================
# SQL 생성
# =============================================================================

def write_sql(out_path: Path, all_data: dict[str, list[dict]]):
    lines: list[str] = []
    p = lines.append

    p("-- ============================================================================")
    p("-- GLOCARE_DB (유학 도메인) 초기 데이터 임포트")
    p("-- 생성일: 2026-04-27")
    p("--")
    p("-- 주의: 0009 마이그레이션 적용 후 실행할 것.")
    p("-- 기존 데이터 보존 (truncate 없음). id 중복 시 conflict.")
    p("-- ============================================================================")
    p("")
    p("begin;")
    p("")

    for sheet, rows in all_data.items():
        if not rows:
            continue
        table = SHEET_TO_TABLE[sheet]
        p(f"-- {sheet} → {table} ({len(rows)} rows)")

        # 첫 row 의 keys 가 컬럼 순서
        cols = list(rows[0].keys())

        for row in rows:
            vals = [sql_quote(row.get(c)) for c in cols]
            p(
                f"insert into public.{table} ({', '.join(cols)}) "
                f"values ({', '.join(vals)});"
            )
        p("")

        # bigserial id 가 충돌 안 나도록 sequence 동기화
        max_id = max((r["id"] for r in rows if r.get("id")), default=0)
        if max_id > 0:
            p(
                f"select setval(pg_get_serial_sequence('public.{table}', 'id'), "
                f"{max_id});"
            )
            p("")

    p("commit;")
    p("")

    out_path.write_text("\n".join(lines), encoding="utf-8")


# =============================================================================
# Main
# =============================================================================

def main():
    if len(sys.argv) < 3:
        print("Usage: python import_glocare_db.py <xlsx-path> <out-sql>")
        sys.exit(1)

    xlsx = Path(sys.argv[1])
    out_sql = Path(sys.argv[2])
    out_sql.parent.mkdir(parents=True, exist_ok=True)

    print(f"Reading: {xlsx}")
    wb = openpyxl.load_workbook(xlsx, data_only=True)

    all_data: dict[str, list[dict]] = {}
    for sheet in HEADER_ROWS.keys():
        if sheet in wb.sheetnames:
            rows = read_sheet(wb, sheet)
            print(f"  {sheet}: {len(rows)} rows")
            all_data[sheet] = rows
        else:
            print(f"  {sheet}: NOT FOUND in workbook")
            all_data[sheet] = []

    print(f"\nWriting SQL: {out_sql}")
    write_sql(out_sql, all_data)
    print("Done!")


if __name__ == "__main__":
    main()
