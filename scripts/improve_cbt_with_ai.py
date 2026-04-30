"""
CBT 해설을 Claude AI 로 재생성.

raw 자료 (CBT 문제 리스트.xlsx + CBT 문제 분석 결과.xlsx) 을 읽어서
각 문제마다 Claude 3.5 Sonnet 호출 → 깔끔하게 정리된 한·베 해설로 변환.

원칙:
  - 보기 해설은 1→5 순서로 정렬
  - 정답은 상세히 설명 (이유·핵심)
  - 오답은 의미 있을 때만 짧게 (순서 정하기 / 단순 매칭 문제는 skip)
  - 베트남어 번역 (한국어 의미 유지)
  - 핵심 용어: TOPIK 3-4 베트남 학습자가 모를 의학·전문용어 위주

진행 상태는 progress.json 에 저장 (중단 후 재개 가능).
출력은 SQL UPDATE 문 (insert 가 아닌 update — 기존 row 갱신용).

사용법:
    # 환경변수 ANTHROPIC_API_KEY 필수
    pip install openpyxl anthropic

    # 5 개만 샘플 테스트
    python scripts/improve_cbt_with_ai.py \\
        "C:/.../CBT 문제 리스트.xlsx" \\
        "C:/.../CBT 문제 분석 결과.xlsx" \\
        --limit 5 \\
        --out supabase/migrations/0013_cbt_ai_sample.sql

    # 전체 (5-8 시간 소요)
    python scripts/improve_cbt_with_ai.py \\
        "C:/.../CBT 문제 리스트.xlsx" \\
        "C:/.../CBT 문제 분석 결과.xlsx" \\
        --out supabase/migrations/0013_cbt_ai_all.sql
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from pathlib import Path

import anthropic
import openpyxl


CHOICE_NUM = "①②③④⑤"


def normalize_chapter(raw):
    if raw is None:
        return "unknown"
    s = str(raw).strip()
    if s == "" or s == "(없음)":
        return "unknown"
    if "그림" in s or "모의고사" in s:
        return "mock"
    m = re.search(r"\d+", s)
    if m:
        return str(int(m.group()))
    return "unknown"


def split_choices(text):
    if text is None:
        return ["", "", "", "", ""]
    s = str(text)
    parts = []
    cur = ""
    cur_num = None
    for ch in s:
        if ch in CHOICE_NUM:
            if cur_num is not None:
                parts.append((cur_num, cur.strip()))
            cur_num = CHOICE_NUM.index(ch) + 1
            cur = ""
        else:
            cur += ch
    if cur_num is not None:
        parts.append((cur_num, cur.strip()))
    parts.sort()
    out = ["", "", "", "", ""]
    for num, val in parts:
        if 1 <= num <= 5:
            out[num - 1] = val
    return out


def sql_str(s):
    if s is None:
        return "NULL"
    return "'" + str(s).replace("\\", "\\\\").replace("'", "''") + "'"


def sql_jsonb(obj):
    if obj is None:
        return "NULL"
    return (
        "'"
        + json.dumps(obj, ensure_ascii=False).replace("'", "''")
        + "'::jsonb"
    )


SYSTEM_PROMPT = """너는 한국 요양보호사 자격시험 CBT 해설 작가야.
대상 학습자는 TOPIK 3-4 급 베트남인 — 한국어 학습 중이라 한국어와 베트남어를 분리해서 보여줘야 함.

각 문제에 대해 다음을 JSON 으로 반환:

1. intent_ko / intent_vi — 이 문제가 무엇을 묻는지 한 줄 요약
   - 한국어와 베트남어 별도로 작성

2. choice_explanations — 보기별 해설. 한국어 블록과 베트남어 블록을 완전히 분리:
   {
     "ko": {"1": "한국어 해설", "2": "...", ...},
     "vi": {"1": "Vietnamese explanation", "2": "...", ...}
   }
   규칙:
   - 정답 보기는 자세히 (왜 옳은지, 핵심 원리)
   - 정답 보기의 해설 맨 앞에 ★ 표시 (한국어·베트남어 둘 다)
   - 오답 보기는 의미 있을 때만 짧게 (잘못된 이유). 순서·매칭 문제처럼 보기별 차이가 무의미하면 해당 키 자체를 생략 (ko·vi 양쪽 다 동일하게)
   - "이건 정답이 아닙니다" 같은 무의미한 해설 금지
   - 1→5 키 순서대로

3. key_terms — TOPIK 3-4 베트남인이 모를 의학·전문 용어 2-5 개
   - 일상 단어 제외 (사람, 식사, 집 등)
   - 전문 용어 우선 (욕창, 흡인, 도뇨, 관장, 인지장애 등)
   - 형식: {"term_ko": "한글 용어", "def_vi": "Vietnamese definition (간단히)"}
   - term_ko 는 한글 그대로, def_vi 는 베트남어로 의미만 짧게

순수 JSON 만 반환 (마크다운 ```json 등 감싸지 말 것)."""


def build_user_prompt(qid, chapter, question, choices, answer_idx, raw_expl, raw_terms):
    return f"""문제 #{qid} (챕터 {chapter})

질문:
{question}

보기:
1) {choices[0]}
2) {choices[1]}
3) {choices[2]}
4) {choices[3]}
5) {choices[4]}

정답: {answer_idx}번

기존 해설 자료 (참고용, 포맷이 들쑥날쑥):
{raw_expl or '(없음)'}

기존 핵심 용어 자료 (참고용):
{raw_terms or '(없음)'}

위 정보를 바탕으로 깔끔히 다시 정리해서 JSON 으로 반환:
{{
  "intent_ko": "...",
  "intent_vi": "...",
  "choice_explanations": {{
    "ko": {{
      "1": "한국어 해설",
      "2": "...",
      "{answer_idx}": "★ 정답 한국어 해설 (자세히)"
    }},
    "vi": {{
      "1": "Vietnamese explanation",
      "2": "...",
      "{answer_idx}": "★ Vietnamese explanation for the correct answer (in detail)"
    }}
  }},
  "key_terms": [
    {{"term_ko": "주보호자", "def_vi": "Người chăm sóc chính"}}
  ]
}}

규칙:
- ko / vi 블록 분리 — 한국어 학습자가 한국어 먼저 읽고 베트남어로 확인하는 흐름
- 정답 ({answer_idx}번) 의 해설은 ★ 로 시작 (한국어·베트남어 둘 다)
- 무의미한 보기는 키 자체 생략 (ko·vi 양쪽 동일하게 — 예: "2" 생략하면 vi 의 "2" 도 생략)
- key_terms 는 [{{term_ko, def_vi}}] 형식만 — term_ko 는 한글 그대로, def_vi 는 짧은 베트남어 정의"""


def call_claude(client, qid, chapter, question, choices, answer_idx, raw_expl, raw_terms, max_retries=3):
    last_err = None
    for attempt in range(max_retries):
        try:
            msg = client.messages.create(
                model="claude-sonnet-4-5",
                max_tokens=2048,
                system=SYSTEM_PROMPT,
                messages=[
                    {
                        "role": "user",
                        "content": build_user_prompt(
                            qid, chapter, question, choices, answer_idx, raw_expl, raw_terms
                        ),
                    }
                ],
            )
            text = msg.content[0].text if msg.content else ""
            jm = re.search(r"\{[\s\S]*\}", text)
            if not jm:
                raise ValueError(f"No JSON in response: {text[:200]}")
            return json.loads(jm.group())
        except (anthropic.APIStatusError, anthropic.APIConnectionError) as e:
            last_err = e
            wait = 5 * (attempt + 1)
            print(f"  ⚠ retry {attempt+1}/{max_retries} after {wait}s: {e}", file=sys.stderr)
            time.sleep(wait)
        except Exception as e:
            last_err = e
            wait = 2 * (attempt + 1)
            print(f"  ⚠ retry {attempt+1}/{max_retries} after {wait}s: {e}", file=sys.stderr)
            time.sleep(wait)
    raise last_err if last_err else RuntimeError("call_claude failed")


def normalize_explanations(choice_expl):
    """입력: {"ko": {"1": ..., ...}, "vi": {"1": ..., ...}}
    출력: 동일한 구조, 단 1→5 키 순서 정렬 + 빈 값 제거.
    """
    if not choice_expl or not isinstance(choice_expl, dict):
        return None
    ko_dict = choice_expl.get("ko") or {}
    vi_dict = choice_expl.get("vi") or {}

    def clean(d):
        out = {}
        for k in sorted(d.keys(), key=lambda x: int(x) if str(x).isdigit() else 99):
            v = d.get(k)
            if isinstance(v, str) and v.strip():
                out[str(k)] = v.strip()
        return out

    ko_clean = clean(ko_dict)
    vi_clean = clean(vi_dict)
    if not ko_clean and not vi_clean:
        return None
    return {"ko": ko_clean, "vi": vi_clean}


def normalize_key_terms(terms):
    """[{term_ko, def_vi}, ...] — 한·베 분리 형식만 유지."""
    if not terms or not isinstance(terms, list):
        return None
    out = []
    for t in terms:
        if not isinstance(t, dict):
            continue
        term_ko = (t.get("term_ko") or "").strip()
        def_vi = (t.get("def_vi") or "").strip()
        if term_ko and def_vi:
            out.append({"term_ko": term_ko, "def_vi": def_vi})
    return out or None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("questions_xlsx")
    ap.add_argument("analysis_xlsx")
    ap.add_argument("--limit", type=int, default=None, help="앞 N 문제만 처리 (테스트용)")
    ap.add_argument("--start-from", type=int, default=0, help="재개 시작 idx")
    ap.add_argument("--out", required=True, help="출력 SQL 파일 경로")
    ap.add_argument("--progress", default="cbt_progress.json", help="진행 상태 저장")
    args = ap.parse_args()

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("ERROR: ANTHROPIC_API_KEY 환경변수 필요", file=sys.stderr)
        sys.exit(1)

    client = anthropic.Anthropic(api_key=api_key)

    # 분석 자료
    wb_a = openpyxl.load_workbook(args.analysis_xlsx, data_only=True)
    ws_a = wb_a.active
    analysis = {}
    for row in ws_a.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        analysis[int(row[0])] = {
            "expl": row[2],
            "terms": row[4] if len(row) > 4 else None,
        }

    # 문제
    wb_q = openpyxl.load_workbook(args.questions_xlsx, data_only=True)
    ws_q = wb_q.active

    questions = []
    for row in ws_q.iter_rows(min_row=2, values_only=True):
        if row[0] is None or row[1] is None or row[3] is None:
            continue
        try:
            qid = int(row[0])
            answer_idx = int(float(row[3]))
        except (ValueError, TypeError):
            continue
        chapter = normalize_chapter(row[4])
        choices = split_choices(row[2])
        if all(c == "" for c in choices):
            continue
        questions.append(
            {
                "id": qid,
                "chapter": chapter,
                "question": row[1],
                "choices": choices,
                "answer_index": answer_idx,
                "raw_expl": (analysis.get(qid) or {}).get("expl"),
                "raw_terms": (analysis.get(qid) or {}).get("terms"),
            }
        )

    if args.limit:
        questions = questions[args.start_from : args.start_from + args.limit]
    elif args.start_from:
        questions = questions[args.start_from :]

    print(f"처리할 문제: {len(questions)} 개", file=sys.stderr)

    # 진행 상태 로드
    progress_path = Path(args.progress)
    done = {}
    if progress_path.exists():
        with progress_path.open(encoding="utf-8") as f:
            done = json.load(f)
        print(f"이미 처리된: {len(done)} 개", file=sys.stderr)

    # 처리
    results = dict(done)
    for i, q in enumerate(questions, 1):
        qid = q["id"]
        if str(qid) in results:
            continue
        try:
            t0 = time.time()
            ai = call_claude(
                client,
                qid,
                q["chapter"],
                q["question"],
                q["choices"],
                q["answer_index"],
                q["raw_expl"],
                q["raw_terms"],
            )
            results[str(qid)] = ai
            dt = time.time() - t0
            print(f"  [{i}/{len(questions)}] #{qid} OK ({dt:.1f}s)", file=sys.stderr)

            # 매 10개마다 progress 저장
            if i % 10 == 0:
                with progress_path.open("w", encoding="utf-8") as f:
                    json.dump(results, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"  [{i}/{len(questions)}] #{qid} FAIL: {e}", file=sys.stderr)
            continue

    # 최종 저장
    with progress_path.open("w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)

    # SQL 생성
    out_path = Path(args.out)
    with out_path.open("w", encoding="utf-8") as out:
        out.write("-- =============================================================================\n")
        out.write("-- 0013_cbt_ai_explanations.sql (auto-generated by improve_cbt_with_ai.py)\n")
        out.write("-- 기존 cbt_questions row 의 intent / explanations / key_terms 를 AI 정리본으로 update.\n")
        out.write("-- =============================================================================\n\n")

        for qid_str, ai in results.items():
            qid = int(qid_str)
            intent_ko = ai.get("intent_ko")
            intent_vi = ai.get("intent_vi")
            ce = normalize_explanations(ai.get("choice_explanations"))
            key_terms = normalize_key_terms(ai.get("key_terms"))

            out.write(
                f"update public.cbt_questions set "
                f"intent_ko = {sql_str(intent_ko)}, "
                f"intent_vi = {sql_str(intent_vi)}, "
                f"choice_explanations = {sql_jsonb(ce)}, "
                f"key_terms = {sql_jsonb(key_terms)} "
                f"where id = {qid};\n"
            )

        out.write(f"\n-- updated: {len(results)} rows\n")

    print(f"\n✅ 완료: {len(results)} 문제 처리됨", file=sys.stderr)
    print(f"   SQL: {out_path}", file=sys.stderr)
    print(f"   진행: {progress_path}", file=sys.stderr)


if __name__ == "__main__":
    main()
