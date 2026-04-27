"""
Glocare 초기 데이터 임포트 스크립트.

엑셀 1개 → SQL 파일 + Review markdown 생성:
  - 교육원 (training_centers) + 강의 일정 (training_classes)
  - 요양원 (care_homes)
  - 교육생 (customers) + 결제 4종 (reservation, welcome_pack, commission, event)
  - 상담 일지 (customer_consultations) — To do list 시트에서

사용법:
    python scripts/import_excel.py <xlsx-path> <out-dir>

주요 정책 (사용자 협의 기준 2026-04-27):
  - 코드는 모두 새로 발급 (CVN/TC/CH + YYMM + 순번)
  - 상태 매핑: 일부는 직접 플래그, 일부는 자동 판정 (`computeCustomerStatus`)
  - 자동판정 후 Excel 단계와 안 맞는 고객 + 모호 케이스는 review markdown 으로
"""

from __future__ import annotations

import datetime
import re
import sys
import uuid
from collections import defaultdict
from pathlib import Path

import openpyxl


# =============================================================================
# 상수
# =============================================================================

# Excel 단계 코드 → DB 플래그 매핑
STATUS_FLAGS: dict[str, dict[str, bool]] = {
    "1-1.단순문의":       {"intake_abandoned": True},
    "1-2.재컨택대기":     {"is_waiting": True},
    "1-3.강의일정확인":   {"class_schedule_confirmation_needed": True},
    "1-4.접수포기":       {"intake_abandoned": True},
    "1-5.유학상담":       {"study_abroad_consultation": True},
    "2-1.교육원발굴":     {"training_center_finding": True},
    "2-3.예약입금완료":   {},  # 자동 판정
    "3-1.교육대기중":     {},
    "3-2.교육중":         {},
    "3-3.교육드랍":       {"training_dropped": True},
    "4-1.시험다시보기":   {},
    "4-2.요양원발굴":     {"care_home_finding": True},
    "4-4.취업대기":       {},
    "4-5.취업포기":       {},  # 사용자 결정 → REVIEW
    "5-1.비자변경대기":   {},
    "5-2.비자완료":       {},
}

# 사용자 review 대상 상태
REVIEW_STATUSES: set[str] = {
    "2-3.예약입금완료",   # 예약금 row 와 일치하는지
    "3-1.교육대기중",     # phase=전 으로 자동 판정되는지
    "3-2.교육중",         # phase=중 으로 자동 판정되는지
    "4-1.시험다시보기",   # 별도 플래그 없음 — 따로 마킹 요청
    "4-4.취업대기",       # 별도 플래그 없음
    "4-5.취업포기",       # 웰컴팩 예약포기 / 근무 종료 — 사용자 결정
    "5-1.비자변경대기",   # 비자 변경일 없으면 안 됨
    "5-2.비자완료",       # 비자 변경일 있어야 됨
}


# =============================================================================
# 유틸
# =============================================================================

def sql_quote(v) -> str:
    """SQL 리터럴 변환."""
    if v is None:
        return "NULL"
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v != v:  # NaN
            return "NULL"
        return str(v) if isinstance(v, int) else str(v)
    if isinstance(v, datetime.datetime):
        return f"'{v.strftime('%Y-%m-%d')}'"
    if isinstance(v, datetime.date):
        return f"'{v.strftime('%Y-%m-%d')}'"
    s = str(v)
    return "'" + s.replace("'", "''") + "'"


def to_date(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, datetime.date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        v = v.strip()
        if not v or v == "-":
            return None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
            try:
                return datetime.datetime.strptime(v, fmt).strftime("%Y-%m-%d")
            except ValueError:
                pass
        return None
    if isinstance(v, (int, float)):
        s = str(int(v))
        if len(s) == 8:
            try:
                return datetime.datetime.strptime(s, "%Y%m%d").strftime("%Y-%m-%d")
            except ValueError:
                return None
    return None


def to_int(v) -> int | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        try:
            n = int(v)
            return n if abs(n) < 1e15 else None
        except (ValueError, OverflowError):
            return None
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        if not s or s == "-":
            return None
        # remove unit suffix (만원, 원, etc.)
        s = re.sub(r"[^\d.\-]", "", s)
        if not s:
            return None
        try:
            return int(float(s))
        except (ValueError, OverflowError):
            return None
    return None


def to_str(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s or s == "-":
        return None
    return s


def normalize_for_match(s) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", "", str(s).strip()).lower()


def gen_code(prefix: str, seq: int) -> str:
    """CVN2604001 형식. seq 는 1부터."""
    yymm = "2604"  # 임포트 기준월 (KST 2026-04)
    return f"{prefix}{yymm}{seq:03d}"


# =============================================================================
# Region 정규화 (시·도 prefix 만 매칭, 나머지는 원본 보존)
# =============================================================================

REGION1_KEYWORDS = [
    ("서울", "서울"),
    ("부산", "부산"),
    ("대구", "대구"),
    ("인천", "인천"),
    ("광주", "광주"),
    ("대전", "대전"),
    ("울산", "울산"),
    ("세종", "세종"),
    ("경기", "경기"),
    ("강원", "강원"),
    ("충북", "충북"),
    ("충청북도", "충북"),
    ("충남", "충남"),
    ("충청남도", "충남"),
    ("전북", "전북"),
    ("전라북도", "전북"),
    ("전남", "전남"),
    ("전라남도", "전남"),
    ("경북", "경북"),
    ("경상북도", "경북"),
    ("경남", "경남"),
    ("경상남도", "경남"),
    ("제주", "제주"),
]


def normalize_region(s) -> str | None:
    raw = to_str(s)
    if not raw:
        return None
    # "서울시", "서울특별시" → "서울"
    for kw, std in REGION1_KEYWORDS:
        if raw.startswith(kw):
            rest = raw[len(kw):].strip()
            # remove trailing 시/도/특별시/광역시 keywords from rest
            rest = re.sub(r"^(특별시|광역시|특별자치도|특별자치시|시|도)\s*", "", rest)
            return f"{std} {rest}".strip() if rest else std
    return raw


# =============================================================================
# Read centers
# =============================================================================

def read_centers(wb) -> list[dict]:
    ws = wb["2.교육원"]
    centers = []
    seq = 0
    for row_idx in range(4, ws.max_row + 1):
        row = list(ws[row_idx])
        name = to_str(row[3].value if len(row) > 3 else None)  # col 4
        if not name:
            continue
        seq += 1
        new_id = str(uuid.uuid4())
        new_code = gen_code("TC", seq)

        # 수강료: 엑셀 입력이 만원/원 혼재라 자동 보정 안 함 — raw 값 보존,
        # 사용자가 사이트에서 한 명씩 보면서 수정
        tuition25 = to_int(row[11].value if len(row) > 11 else None)
        tuition26 = to_int(row[12].value if len(row) > 12 else None)

        contract_text = to_str(row[16].value if len(row) > 16 else None)
        contract_active = bool(contract_text and "완료" in contract_text)

        # 내일배움 카드
        naeil_raw = to_str(row[14].value if len(row) > 14 else None)
        naeil = bool(naeil_raw and "가능" in naeil_raw)

        center = {
            "id": new_id,
            "code": new_code,
            "excel_code": to_str(row[1].value if len(row) > 1 else None),
            "name": name,
            "region": normalize_region(row[2].value if len(row) > 2 else None),
            "address": to_str(row[4].value if len(row) > 4 else None),
            "business_number": to_str(row[5].value if len(row) > 5 else None),
            "director_name": to_str(row[6].value if len(row) > 6 else None),
            "phone": to_str(row[7].value if len(row) > 7 else None),
            "email": to_str(row[8].value if len(row) > 8 else None),
            "bank_name": to_str(row[9].value if len(row) > 9 else None),
            "bank_account": to_str(row[10].value if len(row) > 10 else None),
            "tuition_fee_2025": tuition25,
            "tuition_fee_2026": tuition26,
            "class_hours": to_str(row[13].value if len(row) > 13 else None),
            "naeil_card_eligible": naeil,
            "contract_active": contract_active,
            "deduct_reservation_by_default": True,
            "notes": to_str(row[17].value if len(row) > 17 else None),
            "_row_idx": row_idx,
        }
        centers.append(center)
    return centers


# =============================================================================
# Parse training_classes — 12 monthly columns per center
# =============================================================================

# 월 헤더 → (month) 매핑. 헤더는 "8월 수강정보" 같은 형식.
MONTH_HEADER_PATTERN = re.compile(r"(\d{1,2})\s*월")


def read_classes(wb, centers: list[dict]) -> list[dict]:
    ws = wb["2.교육원"]
    # Top header row 2 = month group label, row 3 = 평일반 개강 / 평반 졸업일 / 야반 개강 / 야반 졸업일
    # Each month block spans 4 columns + 1 blank separator.

    # Detect month blocks: scan row 2 for month markers
    month_blocks: list[dict] = []  # list of {month: int, base_col: int}
    for col in range(20, ws.max_column + 1):
        v = ws.cell(2, col).value
        if v is None:
            continue
        m = MONTH_HEADER_PATTERN.search(str(v))
        if m:
            month_blocks.append({"month": int(m.group(1)), "base_col": col})

    # 첫 블록은 row2 가 비어있는 케이스 — col 20 (헤더 row3 = 평일반 개강) 부터 시작
    # 위 detect 가 col 20 안잡았으면 추가
    if month_blocks and month_blocks[0]["base_col"] > 20:
        # The very first block (col 20-23) might not have a month label in row 2
        first_h3 = ws.cell(3, 20).value
        if first_h3 and "개강" in str(first_h3):
            # Default to month 8 (most common pattern)
            month_blocks.insert(0, {"month": 8, "base_col": 20})

    classes: list[dict] = []
    for center in centers:
        row_idx = center["_row_idx"]
        for block in month_blocks:
            month = block["month"]
            base = block["base_col"]
            # 평일반: base, base+1
            # 야반: base+2, base+3
            wd_start = to_date(ws.cell(row_idx, base).value)
            wd_end = to_date(ws.cell(row_idx, base + 1).value)
            nt_start = to_date(ws.cell(row_idx, base + 2).value)
            nt_end = to_date(ws.cell(row_idx, base + 3).value)

            # 시작일 기반 연도 추정: 8~12월은 2025, 1~7월은 2026
            year_for_month = 2025 if month >= 8 else 2026

            if wd_start:
                classes.append({
                    "id": str(uuid.uuid4()),
                    "training_center_id": center["id"],
                    "year": year_for_month,
                    "month": month,
                    "class_type": "weekday",
                    "start_date": wd_start,
                    "end_date": wd_end,
                })
            if nt_start:
                classes.append({
                    "id": str(uuid.uuid4()),
                    "training_center_id": center["id"],
                    "year": year_for_month,
                    "month": month,
                    "class_type": "night",
                    "start_date": nt_start,
                    "end_date": nt_end,
                })
    return classes


# =============================================================================
# Read homes
# =============================================================================

def read_homes(wb) -> list[dict]:
    ws = wb["3.요양원"]
    homes = []
    seq = 0
    for row_idx in range(4, ws.max_row + 1):
        row = list(ws[row_idx])
        name = to_str(row[2].value if len(row) > 2 else None)  # col 3
        if not name:
            continue
        seq += 1
        homes.append({
            "id": str(uuid.uuid4()),
            "code": gen_code("CH", seq),
            "excel_code": to_str(row[1].value if len(row) > 1 else None),
            "name": name,
            "region": normalize_region(row[3].value if len(row) > 3 else None),
            "address": to_str(row[4].value if len(row) > 4 else None),
            "director_name": to_str(row[5].value if len(row) > 5 else None),
            "phone": to_str(row[6].value if len(row) > 6 else None),
            "contact_person": to_str(row[7].value if len(row) > 7 else None),
            "contact_phone": to_str(row[8].value if len(row) > 8 else None),
            "bed_capacity": to_str(row[9].value if len(row) > 9 else None),
            "partnership_notes": to_str(row[12].value if len(row) > 12 else None),
        })
    return homes


# =============================================================================
# Read customers + payments
# =============================================================================

def find_center_id_by_name(name: str | None, centers: list[dict]) -> str | None:
    if not name:
        return None
    norm = normalize_for_match(name)
    if not norm:
        return None
    # 1) exact normalized match
    for c in centers:
        if normalize_for_match(c["name"]) == norm:
            return c["id"]
    # 2) substring (norm in center name OR center name in norm)
    for c in centers:
        cn = normalize_for_match(c["name"])
        if cn and (norm in cn or cn in norm):
            return c["id"]
    return None


def find_home_id_by_name(name: str | None, homes: list[dict]) -> str | None:
    if not name:
        return None
    norm = normalize_for_match(name)
    if not norm:
        return None
    for h in homes:
        if normalize_for_match(h["name"]) == norm:
            return h["id"]
    for h in homes:
        hn = normalize_for_match(h["name"])
        if hn and (norm in hn or hn in norm):
            return h["id"]
    return None


def find_class_id(
    classes: list[dict],
    center_id: str,
    start_date: str | None,
    class_type: str | None,
) -> str | None:
    if not center_id or not start_date or not class_type:
        return None
    for cls in classes:
        if cls["training_center_id"] != center_id:
            continue
        if cls["class_type"] != class_type:
            continue
        if cls["start_date"] == start_date:
            return cls["id"]
    return None


def parse_class_type(v) -> str | None:
    s = to_str(v)
    if not s:
        return None
    if "주간" in s or "주" in s:
        return "weekday"
    if "야간" in s or "야" in s:
        return "night"
    return None


def parse_product_type(v) -> str | None:
    s = to_str(v)
    if not s:
        return None
    if s in ("교육", "웰컴팩", "교육+웰컴팩"):
        return s
    if "교육" in s and "웰컴팩" in s:
        return "교육+웰컴팩"
    if "웰컴팩" in s:
        return "웰컴팩"
    if "교육" in s:
        return "교육"
    return None


def parse_gender(v) -> str | None:
    s = to_str(v)
    if not s:
        return None
    if "여" in s:
        return "여"
    if "남" in s:
        return "남"
    return None


def parse_desired_time(v) -> str | None:
    s = to_str(v)
    if not s:
        return None
    if "주간" in s:
        return "주간"
    if "야간" in s:
        return "야간"
    return None


def parse_topik(v) -> str | None:
    s = to_str(v)
    if not s:
        return None
    # patterns like "TOPIK 3급", "3급", "3", etc.
    m = re.search(r"(\d)", s)
    if m:
        return f"{m.group(1)}급"
    return s


def read_customers(wb, centers: list[dict], homes: list[dict], classes: list[dict]) -> tuple[list[dict], list[dict], list[dict], list[dict], list[dict], list[dict]]:
    """
    Returns: customers, reservation_payments, welcome_pack_payments,
             commission_payments, event_payments, review_entries
    """
    ws = wb["1.고객관리"]
    customers = []
    reservations = []
    welcome_packs = []
    commissions = []
    events = []
    reviews = []  # {code, name, status, reason, ...}

    seq = 0
    for row_idx in range(5, ws.max_row + 1):
        row = list(ws[row_idx])
        excel_code = to_str(row[1].value if len(row) > 1 else None)  # col 2
        if not excel_code:
            continue

        seq += 1
        customer_id = str(uuid.uuid4())
        new_code = gen_code("CVN", seq)

        excel_status = to_str(row[2].value if len(row) > 2 else None)  # col 3
        flags_extra = STATUS_FLAGS.get(excel_status, {}) if excel_status else {}

        name_vi = to_str(row[4].value if len(row) > 4 else None)
        name_kr = to_str(row[5].value if len(row) > 5 else None)
        address = to_str(row[6].value if len(row) > 6 else None)
        gender = parse_gender(row[7].value if len(row) > 7 else None)
        birth_year = to_int(row[8].value if len(row) > 8 else None)
        phone = to_str(row[10].value if len(row) > 10 else None)  # col 11
        visa = to_str(row[11].value if len(row) > 11 else None)
        topik = parse_topik(row[12].value if len(row) > 12 else None)
        stay_remaining = to_str(row[13].value if len(row) > 13 else None)

        # 희망 조건
        desired_period_raw = row[15].value if len(row) > 15 else None  # col 16
        if isinstance(desired_period_raw, datetime.datetime):
            desired_period = desired_period_raw.strftime("%Y-%m")
        else:
            desired_period = to_str(desired_period_raw)
        desired_time = parse_desired_time(row[16].value if len(row) > 16 else None)
        desired_region = normalize_region(row[17].value if len(row) > 17 else None)

        # 교육원 매칭 (col 21 = 교육원 이름)
        center_name = to_str(row[20].value if len(row) > 20 else None)
        training_center_id = find_center_id_by_name(center_name, centers)

        class_start = to_date(row[22].value if len(row) > 22 else None)  # col 23
        class_type = parse_class_type(row[23].value if len(row) > 23 else None)
        training_class_id = (
            find_class_id(classes, training_center_id, class_start, class_type)
            if training_center_id
            else None
        )
        # 강의 시작/종료는 customer 가 갖고있긴 하지만 (denorm) 우리 폼은 training_class 에서 끌어옴
        # 매칭 안 된 경우에도 raw class_start 는 보존 (참고용)
        class_end = None
        if training_class_id:
            for cls in classes:
                if cls["id"] == training_class_id:
                    class_end = cls["end_date"]
                    break

        # 요양원 매칭
        home_name = to_str(row[28].value if len(row) > 28 else None)  # col 29
        care_home_id = find_home_id_by_name(home_name, homes)

        work_start = to_date(row[30].value if len(row) > 30 else None)  # col 31
        visa_change = to_date(row[31].value if len(row) > 31 else None)  # col 32

        # 상품 (col 34)
        product_type = parse_product_type(row[33].value if len(row) > 33 else None)

        # ===== 결제: 예약금 (col 36~41) =====
        res_amount = to_int(row[36].value if len(row) > 36 else None)
        res_pay_date = to_date(row[37].value if len(row) > 37 else None)
        res_refund_amount = to_int(row[38].value if len(row) > 38 else None)
        res_refund_date = to_date(row[39].value if len(row) > 39 else None)
        if (res_amount and res_amount > 0) or res_pay_date:
            reservations.append({
                "customer_id": customer_id,
                "amount": res_amount or 0,
                "payment_date": res_pay_date,
                "refund_amount": res_refund_amount or 0,
                "refund_date": res_refund_date,
                "refund_reason": None,  # 엑셀 자유텍스트라 매핑 어려움
            })

        # ===== 결제: 소개비 (col 42~44) — 사용자 정책: "다 넣어줘" =====
        # 신규 commission_payments 스키마: settlement_month, total_amount, deduction_amount
        comm_total = to_int(row[41].value if len(row) > 41 else None)  # col 42
        comm_received = to_int(row[42].value if len(row) > 42 else None)  # col 43
        comm_paid_date = to_date(row[43].value if len(row) > 43 else None)  # col 44
        if comm_paid_date and training_center_id and (comm_total or comm_received):
            settlement_month = comm_paid_date[:7] + "-01"
            total_amount = comm_total or 0
            received_amount = comm_received or 0
            deduction = max(0, total_amount - received_amount)
            commissions.append({
                "customer_id": customer_id,
                "training_center_id": training_center_id,
                "settlement_month": settlement_month,
                "total_amount": total_amount,
                "deduction_amount": deduction,
                "completed_at": comm_paid_date,
            })

        # ===== 결제: 이벤트 (col 45~48) =====
        ev_type = to_str(row[44].value if len(row) > 44 else None)  # col 45
        ev_gift = to_str(row[45].value if len(row) > 45 else None)  # col 46
        ev_amount = to_int(row[46].value if len(row) > 46 else None)  # col 47
        ev_given_date = to_date(row[47].value if len(row) > 47 else None)  # col 48
        if ev_type or ev_amount or ev_given_date:
            events.append({
                "customer_id": customer_id,
                "event_type": ev_type or "기타",
                "amount": ev_amount or 0,
                "gift_type": ev_gift,
                "friend_customer_id": None,
                "gift_given": bool(ev_given_date),
                "gift_given_date": ev_given_date,
            })

        # ===== 결제: 웰컴팩 (col 49~56) =====
        wp_total = to_int(row[48].value if len(row) > 48 else None)  # col 49
        wp_discount = to_int(row[49].value if len(row) > 49 else None)  # col 50
        wp_res_amount = to_int(row[51].value if len(row) > 51 else None)  # col 52
        wp_res_date = to_date(row[52].value if len(row) > 52 else None)  # col 53
        wp_int_amount = to_int(row[53].value if len(row) > 53 else None)  # col 54
        wp_int_date = to_date(row[54].value if len(row) > 54 else None)  # col 55
        wp_bal_amount = to_int(row[55].value if len(row) > 55 else None)  # col 56
        if any([wp_total, wp_discount, wp_res_amount, wp_res_date, wp_int_amount, wp_int_date, wp_bal_amount]):
            welcome_packs.append({
                "customer_id": customer_id,
                "total_price": wp_total or 1500000,
                "discount_amount": wp_discount or 0,
                "reservation_amount": wp_res_amount or 0,
                "reservation_date": wp_res_date,
                "interim_amount": wp_int_amount or 0,
                "interim_date": wp_int_date,
                "balance_amount": wp_bal_amount or 0,
                "balance_date": None,
                "sales_reported": False,
                "sales_reported_date": None,
            })

        # 종료 사유 (4-5.취업포기 — 사용자가 추후 결정)
        termination_reason = None

        # 플래그 기본값
        flags = {
            "intake_abandoned": False,
            "study_abroad_consultation": False,
            "training_center_finding": False,
            "class_schedule_confirmation_needed": False,
            "training_reservation_abandoned": False,
            "certificate_acquired": False,
            "training_dropped": False,
            "welcome_pack_abandoned": False,
            "care_home_finding": False,
            "resume_sent": False,
            "interview_passed": False,
        }
        flags.update({k: v for k, v in flags_extra.items() if k in flags})

        # is_waiting / 대기 메모
        is_waiting = bool(flags_extra.get("is_waiting"))

        customer = {
            "id": customer_id,
            "code": new_code,
            "excel_code": excel_code,
            "name_vi": name_vi,
            "name_kr": name_kr,
            "phone": phone,
            "address": address,
            "gender": gender,
            "birth_year": birth_year,
            "visa_type": visa,
            "topik_level": topik,
            "stay_remaining": stay_remaining,
            "desired_period": desired_period,
            "desired_time": desired_time,
            "desired_region": desired_region,
            "training_center_id": training_center_id,
            "training_class_id": training_class_id,
            "care_home_id": care_home_id,
            "class_start_date": class_start if not training_class_id else (
                # use class's start (denorm sync)
                next((c["start_date"] for c in classes if c["id"] == training_class_id), class_start)
            ),
            "class_end_date": class_end,
            "work_start_date": work_start,
            "work_end_date": None,
            "visa_change_date": visa_change,
            "interview_date": None,
            "product_type": product_type,
            "is_waiting": is_waiting,
            "recontact_date": None,
            "waiting_memo": None,
            "termination_reason": termination_reason,
            "legacy_status": excel_status,
            "_excel_row": row_idx,
            "_excel_status": excel_status,
            "_excel_center_name": center_name,
            "_excel_home_name": home_name,
            "flags": flags,
        }
        customers.append(customer)

        # ============== Review 작성 ==============
        review_reason = []
        if excel_status in REVIEW_STATUSES:
            review_reason.append(f"상태={excel_status}")
        if center_name and not training_center_id:
            review_reason.append(f"교육원 매칭 실패: '{center_name}'")
        if home_name and not care_home_id:
            review_reason.append(f"요양원 매칭 실패: '{home_name}'")
        if class_start and class_type and training_center_id and not training_class_id:
            review_reason.append(
                f"강의 매칭 실패: {class_start} {class_type} (교육원 일정에 등록 필요)"
            )
        if excel_status == "5-1.비자변경대기" and not work_start:
            review_reason.append("비자변경대기인데 근무시작일 없음")
        if excel_status == "5-2.비자완료" and not visa_change:
            review_reason.append("비자완료인데 비자변경일 없음")

        if review_reason:
            reviews.append({
                "code": new_code,
                "excel_code": excel_code,
                "name": name_kr or name_vi or "(이름 없음)",
                "status": excel_status,
                "reasons": review_reason,
            })

    return customers, reservations, welcome_packs, commissions, events, reviews


# =============================================================================
# Read To do list → consultations
# =============================================================================

def read_consultations(wb, customers: list[dict]) -> list[dict]:
    ws = wb["To do list"]
    # 코드 → customer_id 맵
    code_to_id = {c["excel_code"]: c["id"] for c in customers if c["excel_code"]}

    consultations = []
    for row_idx in range(2, ws.max_row + 1):
        row = list(ws[row_idx])
        excel_code = to_str(row[1].value if len(row) > 1 else None)  # col 2
        if not excel_code or excel_code not in code_to_id:
            continue
        # 작성일자 col 1
        created = row[0].value if len(row) > 0 else None
        type_str = to_str(row[3].value if len(row) > 3 else None)  # col 4
        original = to_str(row[4].value if len(row) > 4 else None)  # col 5 (원본 베트남어/한국어)
        translated = to_str(row[5].value if len(row) > 5 else None)  # col 6 (한글 번역)
        result = to_str(row[6].value if len(row) > 6 else None)  # col 7 (결과)

        # 컨텐츠 만들기 — 베트남어 + 한글 번역 + 결과
        parts = []
        if type_str:
            parts.append(f"[{type_str}]")
        if original:
            parts.append(original)
        # content_kr: 한글 번역 우선, 없으면 결과
        # content_vi: original 이 베트남어 추정이면 거기에
        # 단순화: content_kr = translated or result, content_vi = original 베트남어 패턴이면 거기

        is_vietnamese = bool(original and re.search(r"[\u1ea0-\u1ef9\u00C0-\u1EF9]", original))

        content_vi = original if is_vietnamese else None
        content_kr = translated if translated else (original if not is_vietnamese else None)
        if result:
            content_kr = (content_kr + "\n\n[결과] " + result) if content_kr else f"[결과] {result}"

        # consultation_type 추론
        ctype = "training_center"
        if type_str:
            if "요양원" in type_str:
                ctype = "care_home"
            elif "교육원" in type_str or "시간표" in type_str or "강의" in type_str:
                ctype = "training_center"

        consultations.append({
            "id": str(uuid.uuid4()),
            "customer_id": code_to_id[excel_code],
            "consultation_type": ctype,
            "content_vi": content_vi,
            "content_kr": content_kr,
            "tags": [],
            "created_at": created.strftime("%Y-%m-%d %H:%M:%S") if isinstance(created, datetime.datetime) else None,
        })
    return consultations


# =============================================================================
# SQL 생성
# =============================================================================

def write_sql(
    out_path: Path,
    centers: list[dict],
    classes: list[dict],
    homes: list[dict],
    customers: list[dict],
    reservations: list[dict],
    welcome_packs: list[dict],
    commissions: list[dict],
    events: list[dict],
    consultations: list[dict],
):
    lines: list[str] = []
    p = lines.append

    p("-- ============================================================================")
    p("-- Glocare 초기 데이터 임포트 (생성일: 2026-04-27)")
    p("--")
    p("-- 생성 통계:")
    p(f"--   training_centers: {len(centers)}")
    p(f"--   training_classes: {len(classes)}")
    p(f"--   care_homes: {len(homes)}")
    p(f"--   customers: {len(customers)}")
    p(f"--   reservation_payments: {len(reservations)}")
    p(f"--   welcome_pack_payments: {len(welcome_packs)}")
    p(f"--   commission_payments: {len(commissions)}")
    p(f"--   event_payments: {len(events)}")
    p(f"--   customer_consultations: {len(consultations)}")
    p("-- ============================================================================")
    p("")
    p("begin;")
    p("")
    p("-- 기존 데이터 클린 (의존성 순서). auth.users 는 안 건드림.")
    p("truncate table public.customer_consultations,")
    p("                public.reservation_payments,")
    p("                public.welcome_pack_payments,")
    p("                public.commission_payments,")
    p("                public.event_payments,")
    p("                public.sms_messages,")
    p("                public.customer_statuses,")
    p("                public.customers,")
    p("                public.training_classes,")
    p("                public.care_homes,")
    p("                public.training_centers")
    p("  restart identity cascade;")
    p("")

    # ========== training_centers ==========
    p("-- training_centers")
    cols = [
        "id", "code", "name", "region", "address", "business_number",
        "director_name", "phone", "email", "bank_name", "bank_account",
        "tuition_fee_2025", "tuition_fee_2026", "class_hours",
        "naeil_card_eligible", "contract_active", "deduct_reservation_by_default",
        "notes",
    ]
    for c in centers:
        vals = [sql_quote(c[k]) for k in cols]
        p(f"insert into public.training_centers ({', '.join(cols)}) values ({', '.join(vals)});")
    p("")

    # ========== training_classes ==========
    p("-- training_classes")
    cls_cols = ["id", "training_center_id", "year", "month", "class_type", "start_date", "end_date"]
    for cls in classes:
        vals = [sql_quote(cls[k]) for k in cls_cols]
        p(f"insert into public.training_classes ({', '.join(cls_cols)}) values ({', '.join(vals)});")
    p("")

    # ========== care_homes ==========
    p("-- care_homes")
    home_cols = [
        "id", "code", "name", "region", "address", "director_name", "phone",
        "contact_person", "contact_phone", "bed_capacity", "partnership_notes",
    ]
    for h in homes:
        vals = [sql_quote(h[k]) for k in home_cols]
        p(f"insert into public.care_homes ({', '.join(home_cols)}) values ({', '.join(vals)});")
    p("")

    # ========== customers ==========
    p("-- customers")
    cust_cols = [
        "id", "code", "name_vi", "name_kr", "phone", "address", "gender",
        "birth_year", "visa_type", "topik_level", "stay_remaining",
        "desired_period", "desired_time", "desired_region",
        "training_center_id", "training_class_id", "care_home_id",
        "class_start_date", "class_end_date", "work_start_date", "work_end_date",
        "visa_change_date", "interview_date", "product_type",
        "is_waiting", "recontact_date", "waiting_memo",
        "termination_reason", "legacy_status",
    ]
    for c in customers:
        row = {k: c.get(k) for k in cust_cols}
        vals = [sql_quote(row[k]) for k in cust_cols]
        p(f"insert into public.customers ({', '.join(cust_cols)}) values ({', '.join(vals)});")
    p("")

    # customer_statuses 는 트리거로 생성됨. 거기에 플래그 update.
    p("-- customer_statuses (트리거로 default row 생성됨 — 플래그만 update)")
    flag_keys = [
        "intake_abandoned", "study_abroad_consultation", "training_center_finding",
        "class_schedule_confirmation_needed", "training_reservation_abandoned",
        "certificate_acquired", "training_dropped", "welcome_pack_abandoned",
        "care_home_finding", "resume_sent", "interview_passed",
    ]
    for c in customers:
        flags = c["flags"]
        non_default = {k: v for k, v in flags.items() if v}
        if non_default:
            sets = ", ".join(f"{k} = {sql_quote(v)}" for k, v in non_default.items())
            p(f"update public.customer_statuses set {sets} where customer_id = {sql_quote(c['id'])};")
    p("")

    # ========== reservation_payments ==========
    p("-- reservation_payments")
    res_cols = [
        "customer_id", "amount", "payment_date",
        "refund_amount", "refund_date", "refund_reason",
    ]
    for r in reservations:
        vals = [sql_quote(r[k]) for k in res_cols]
        p(f"insert into public.reservation_payments ({', '.join(res_cols)}) values ({', '.join(vals)});")
    p("")

    # ========== welcome_pack_payments ==========
    p("-- welcome_pack_payments")
    wp_cols = [
        "customer_id", "total_price", "discount_amount",
        "reservation_amount", "reservation_date",
        "interim_amount", "interim_date",
        "balance_amount", "balance_date",
        "sales_reported", "sales_reported_date",
    ]
    for w in welcome_packs:
        vals = [sql_quote(w[k]) for k in wp_cols]
        p(f"insert into public.welcome_pack_payments ({', '.join(wp_cols)}) values ({', '.join(vals)});")
    p("")

    # ========== commission_payments ==========
    p("-- commission_payments")
    com_cols = [
        "customer_id", "training_center_id", "settlement_month",
        "total_amount", "deduction_amount", "completed_at",
    ]
    for c in commissions:
        vals = [sql_quote(c[k]) for k in com_cols]
        p(f"insert into public.commission_payments ({', '.join(com_cols)}) values ({', '.join(vals)});")
    p("")

    # ========== event_payments ==========
    p("-- event_payments")
    ev_cols = [
        "customer_id", "event_type", "amount", "gift_type",
        "friend_customer_id", "gift_given", "gift_given_date",
    ]
    for e in events:
        vals = [sql_quote(e[k]) for k in ev_cols]
        p(f"insert into public.event_payments ({', '.join(ev_cols)}) values ({', '.join(vals)});")
    p("")

    # ========== consultations ==========
    p("-- customer_consultations")
    cons_cols = [
        "id", "customer_id", "consultation_type",
        "content_vi", "content_kr", "tags", "created_at",
    ]
    for cons in consultations:
        # tags array
        tags_sql = "ARRAY[]::text[]" if not cons["tags"] else (
            "ARRAY[" + ", ".join(sql_quote(t) for t in cons["tags"]) + "]::text[]"
        )
        vals = [
            sql_quote(cons["id"]),
            sql_quote(cons["customer_id"]),
            sql_quote(cons["consultation_type"]),
            sql_quote(cons["content_vi"]),
            sql_quote(cons["content_kr"]),
            tags_sql,
            sql_quote(cons["created_at"]) if cons["created_at"] else "now()",
        ]
        p(f"insert into public.customer_consultations ({', '.join(cons_cols)}) values ({', '.join(vals)});")
    p("")

    p("commit;")
    p("")

    out_path.write_text("\n".join(lines), encoding="utf-8")


def write_review(out_path: Path, reviews: list[dict], customers: list[dict]):
    lines: list[str] = []
    p = lines.append

    # 그룹: 4-5 (최우선), 그 다음 5-1/5-2, 그 다음 매칭 실패류, 그 다음 자동 판정 검증
    groups = defaultdict(list)
    for r in reviews:
        if r["status"] == "4-5.취업포기":
            groups["a_4-5_취업포기"].append(r)
        elif r["status"] in ("5-1.비자변경대기", "5-2.비자완료"):
            groups["b_5-x_비자관련"].append(r)
        elif r["status"] in ("2-3.예약입금완료", "3-1.교육대기중", "3-2.교육중"):
            groups["c_자동판정검증"].append(r)
        elif r["status"] in ("4-1.시험다시보기", "4-4.취업대기"):
            groups["d_별도플래그없음"].append(r)
        else:
            groups["e_매칭실패등"].append(r)

    p("# 임포트 후 사용자 검토 필요 리스트")
    p("")
    p(f"총 검토 대상: **{len(reviews)}명** / 전체 {len(customers)}명")
    p("")
    p("> 임포트 후 사이트에서 한 명씩 보면서 결정/수정.")
    p("")

    titles = {
        "a_4-5_취업포기": "## A. 4-5.취업포기 → 웰컴팩 예약포기 / 근무 종료 결정 필요",
        "b_5-x_비자관련": "## B. 비자 관련 (5-1/5-2) → 근무시작일·비자변경일 보강 필요",
        "c_자동판정검증": "## C. 자동 판정 검증 (2-3/3-1/3-2)",
        "d_별도플래그없음": "## D. 별도 플래그 없음 (4-1/4-4) — 진행 단계 그대로 검토",
        "e_매칭실패등": "## E. 매칭 실패 또는 기타 — 수기 정리 필요",
    }

    for k in ["a_4-5_취업포기", "b_5-x_비자관련", "c_자동판정검증", "d_별도플래그없음", "e_매칭실패등"]:
        items = groups.get(k, [])
        if not items:
            continue
        p(titles[k])
        p("")
        p(f"({len(items)}명)")
        p("")
        p("| 코드 | 이름 | Excel 상태 | 사유 |")
        p("| --- | --- | --- | --- |")
        for r in sorted(items, key=lambda x: x["code"]):
            reasons = " / ".join(r["reasons"])
            p(f"| {r['code']} | {r['name']} | {r['status'] or '—'} | {reasons} |")
        p("")

    out_path.write_text("\n".join(lines), encoding="utf-8")


# =============================================================================
# Main
# =============================================================================

def main():
    if len(sys.argv) < 3:
        print("Usage: python import_excel.py <xlsx-path> <out-dir>")
        sys.exit(1)

    xlsx_path = Path(sys.argv[1])
    out_dir = Path(sys.argv[2])
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"Reading: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    print("Parsing centers...")
    centers = read_centers(wb)
    print(f"  centers: {len(centers)}")

    print("Parsing classes...")
    classes = read_classes(wb, centers)
    print(f"  classes: {len(classes)}")

    print("Parsing care homes...")
    homes = read_homes(wb)
    print(f"  homes: {len(homes)}")

    print("Parsing customers...")
    customers, reservations, welcome_packs, commissions, events, reviews = read_customers(
        wb, centers, homes, classes
    )
    print(f"  customers: {len(customers)}")
    print(f"  reservations: {len(reservations)}")
    print(f"  welcome_packs: {len(welcome_packs)}")
    print(f"  commissions: {len(commissions)}")
    print(f"  events: {len(events)}")
    print(f"  reviews: {len(reviews)}")

    print("Parsing consultations from To do list...")
    consultations = read_consultations(wb, customers)
    print(f"  consultations: {len(consultations)}")

    sql_path = out_dir / "data_import.sql"
    review_path = out_dir / "import-review-needed.md"

    print(f"\nWriting SQL: {sql_path}")
    write_sql(sql_path, centers, classes, homes, customers,
              reservations, welcome_packs, commissions, events, consultations)

    print(f"Writing review: {review_path}")
    write_review(review_path, reviews, customers)

    print("\nDone!")


if __name__ == "__main__":
    main()
