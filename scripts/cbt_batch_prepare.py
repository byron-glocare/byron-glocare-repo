"""
Claude Max 채팅창용 batch 프롬프트 파일 생성.

API 사용 없이 Claude.ai (Max 구독) 에 paste 해서 처리 → 결과 JSON 저장 →
나중에 cbt_batch_to_sql.py 로 SQL 변환.

사용법:
    python scripts/cbt_batch_prepare.py \\
        "C:/.../CBT 문제 리스트.xlsx" \\
        "C:/.../CBT 문제 분석 결과.xlsx" \\
        --batch-size 30 \\
        --out scripts/cbt_batches

→ scripts/cbt_batches/batch_001.txt, batch_002.txt, ... 생성
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from pathlib import Path

import openpyxl


CHOICE_NUM = "①②③④⑤"


def normalize_chapter(raw):
    if raw is None:
        return "unknown"
    s = str(raw).strip()
    if s == "" or s == "(없음)":
        return "unknown"
    if "그림" in s or "모의고사" in s:
        return "mock"
    m = re.search(r"\d+", s)
    if m:
        return str(int(m.group()))
    return "unknown"


def split_choices(text):
    if text is None:
        return ["", "", "", "", ""]
    s = str(text)
    parts = []
    cur = ""
    cur_num = None
    for ch in s:
        if ch in CHOICE_NUM:
            if cur_num is not None:
                parts.append((cur_num, cur.strip()))
            cur_num = CHOICE_NUM.index(ch) + 1
            cur = ""
        else:
            cur += ch
    if cur_num is not None:
        parts.append((cur_num, cur.strip()))
    parts.sort()
    out = ["", "", "", "", ""]
    for num, val in parts:
        if 1 <= num <= 5:
            out[num - 1] = val
    return out


SYSTEM_PROMPT = """너는 한국 요양보호사 자격시험 CBT 해설 작가야.
대상 학습자는 TOPIK 3-4 급 베트남인 — 한국어 학습 중이라 한국어와 베트남어를 분리해서 보여줘야 함.

각 문제에 대해 다음 형식으로 JSON 반환:
- intent_ko / intent_vi — 한 줄 요약
- choice_explanations — { "ko": {1: ...}, "vi": {1: ...} } — ko/vi 블록 분리, 1→5 순서
  - 정답 보기는 자세히 (왜 옳은지, 핵심 원리), 해설 맨 앞에 ★
  - 오답 보기는 의미 있을 때만 짧게 (잘못된 이유)
  - 순서·매칭 문제처럼 보기별 차이가 무의미하면 해당 키 자체를 생략 (ko·vi 양쪽 동일)
  - "이건 정답이 아닙니다" 같은 무의미한 해설 금지
- key_terms — [{ "term_ko": "한글 용어", "def_vi": "Vietnamese definition" }] 형식
  - TOPIK 3-4 가 모를 의학·전문 용어 2-5 개
  - 일상 단어 제외, 전문 용어 우선

여러 문제를 한 번에 처리할 때는 array 로 묶어서 반환."""


def build_batch_prompt(batch_questions):
    parts = [SYSTEM_PROMPT, ""]
    parts.append(f"다음 {len(batch_questions)} 문제를 모두 처리해서 array 로 반환:")
    parts.append("")
    for q in batch_questions:
        parts.append(f"=== 문제 ID {q['id']} (챕터 {q['chapter']}) ===")
        parts.append("질문:")
        parts.append(q["question"])
        parts.append("")
        parts.append("보기:")
        for i, c in enumerate(q["choices"], 1):
            parts.append(f"{i}) {c}")
        parts.append("")
        parts.append(f"정답: {q['answer_index']}번")
        if q["raw_expl"]:
            parts.append("")
            parts.append("기존 해설 자료 (참고용):")
            parts.append(str(q["raw_expl"]))
        if q["raw_terms"]:
            parts.append("")
            parts.append("기존 핵심 용어 자료 (참고용):")
            parts.append(str(q["raw_terms"]))
        parts.append("")
        parts.append("---")
        parts.append("")

    parts.append("위 모든 문제를 처리해서 다음 형식의 JSON array 만 반환 (마크다운 ```json 등 감싸지 말 것):")
    parts.append("")
    parts.append("```")
    parts.append('[')
    parts.append('  {')
    parts.append('    "id": 1740,')
    parts.append('    "intent_ko": "...",')
    parts.append('    "intent_vi": "...",')
    parts.append('    "choice_explanations": {')
    parts.append('      "ko": { "1": "...", "2": "★ ..." },')
    parts.append('      "vi": { "1": "...", "2": "★ ..." }')
    parts.append('    },')
    parts.append('    "key_terms": [{ "term_ko": "주보호자", "def_vi": "Người chăm sóc chính" }]')
    parts.append('  },')
    parts.append('  ... (각 id 마다 한 객체)')
    parts.append(']')
    parts.append("```")
    parts.append("")
    parts.append("⚠️ 정답에 해당하는 보기의 해설은 반드시 ★ 로 시작 (한국어·베트남어 둘 다)")
    parts.append("⚠️ 무의미한 보기는 키 자체 생략 (ko·vi 양쪽 동일하게)")
    parts.append("⚠️ 한국어와 베트남어 블록 완전 분리")

    return "\n".join(parts)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("questions_xlsx")
    ap.add_argument("analysis_xlsx")
    ap.add_argument("--batch-size", type=int, default=30, help="배치당 문제 수 (기본 30)")
    ap.add_argument("--out", default="scripts/cbt_batches", help="출력 디렉터리")
    ap.add_argument("--start-id", type=int, default=None, help="이 ID 부터 처리 (재개용)")
    args = ap.parse_args()

    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    # 분석 자료
    wb_a = openpyxl.load_workbook(args.analysis_xlsx, data_only=True)
    ws_a = wb_a.active
    analysis = {}
    for row in ws_a.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        analysis[int(row[0])] = {
            "expl": row[2],
            "terms": row[4] if len(row) > 4 else None,
        }

    # 문제
    wb_q = openpyxl.load_workbook(args.questions_xlsx, data_only=True)
    ws_q = wb_q.active

    questions = []
    for row in ws_q.iter_rows(min_row=2, values_only=True):
        if row[0] is None or row[1] is None or row[3] is None:
            continue
        try:
            qid = int(row[0])
            answer_idx = int(float(row[3]))
        except (ValueError, TypeError):
            continue
        if args.start_id and qid < args.start_id:
            continue
        chapter = normalize_chapter(row[4])
        choices = split_choices(row[2])
        if all(c == "" for c in choices):
            continue
        questions.append(
            {
                "id": qid,
                "chapter": chapter,
                "question": row[1],
                "choices": choices,
                "answer_index": answer_idx,
                "raw_expl": (analysis.get(qid) or {}).get("expl"),
                "raw_terms": (analysis.get(qid) or {}).get("terms"),
            }
        )

    print(f"총 문제: {len(questions)}", file=sys.stderr)

    # 배치 분할
    n = args.batch_size
    batches = [questions[i : i + n] for i in range(0, len(questions), n)]
    print(f"배치 수: {len(batches)} (각 ~{n} 문제)", file=sys.stderr)

    # 각 배치 → 파일 출력
    index = []
    for bi, batch in enumerate(batches, 1):
        path = out_dir / f"batch_{bi:03d}.txt"
        prompt = build_batch_prompt(batch)
        with path.open("w", encoding="utf-8") as f:
            f.write(prompt)
        ids = [q["id"] for q in batch]
        index.append({"batch": bi, "file": path.name, "ids": ids, "count": len(ids)})
        print(f"  {path.name}: ids {ids[0]}~{ids[-1]} ({len(ids)} 문제)", file=sys.stderr)

    # 인덱스 파일
    with (out_dir / "index.json").open("w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False, indent=2)

    # README
    readme = f"""# CBT Batch Workflow (Claude Max 채팅용)

## 흐름
1. 각 batch_NNN.txt 파일을 메모장/VSCode 로 연다
2. 전체 선택 (Ctrl+A) → 복사 → Claude.ai (Max) 채팅창에 붙여넣기 → 전송
3. Claude 가 JSON array 로 응답
4. 응답 JSON 만 복사해서 batch_NNN_response.json 으로 저장
5. 모든 batch 완료 후 cbt_batch_to_sql.py 로 SQL 변환

## 진행 상황
- 총 {len(questions)} 문제, {len(batches)} 배치
- 배치당 약 {n} 문제

## 다음 단계
모든 batch_NNN_response.json 저장 완료 후:
```
python scripts/cbt_batch_to_sql.py {args.out} --out supabase/migrations/0013_cbt_ai_all.sql
```
"""
    (out_dir / "README.md").write_text(readme, encoding="utf-8")

    print(f"\n✅ {out_dir}/ 에 {len(batches)} 개 배치 파일 생성", file=sys.stderr)
    print(f"   index.json + README.md 도 같이", file=sys.stderr)


if __name__ == "__main__":
    main()
