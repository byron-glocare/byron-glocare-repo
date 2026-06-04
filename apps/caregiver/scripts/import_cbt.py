"""
CBT 문제 + 분석 자료 → SQL INSERT 변환.

사용법:
    python scripts/import_cbt.py \
        "C:/Users/kajka/Downloads/CBT 문제 리스트.xlsx" \
        "C:/Users/kajka/Downloads/CBT 문제 분석 결과.xlsx" \
        > supabase/migrations/0012_cbt_questions_data.sql

처리:
  - 챕터 normalize: '챕터1', 'CHAPTER 01', '챱터5', '팹터7' 등 → '1', '5', '7'
  - '그림 중심 실전모의고사' → 'mock'
  - 정답 1.0~5.0 → 정수 1~5
  - 보기 텍스트 (① ② ③ ④ ⑤) → JSON array
  - 분석 자료 (문제 의도 / 보기 해설 / 핵심용어) join
  - 한·베 분리해서 jsonb 로
"""
from __future__ import annotations

import re
import sys
import json
from pathlib import Path

import openpyxl


def normalize_chapter(raw):
    """챕터 값 정리. 'CHAPTER 01' / '챕터1' / '챱터5' / '팹터7' → '1','5','7'."""
    if raw is None:
        return "unknown"
    s = str(raw).strip()
    if s == "" or s == "(없음)":
        return "unknown"
    if "그림" in s or "모의고사" in s:
        return "mock"
    # Extract number
    m = re.search(r"\d+", s)
    if m:
        return str(int(m.group()))
    return "unknown"


CHOICE_NUM = "①②③④⑤"


def split_choices(text):
    """'① 보기1 ② 보기2 ...' → ['보기1', '보기2', ...]"""
    if text is None:
        return []
    s = str(text)
    # Split by each ① ② ③ ④ ⑤
    parts = []
    cur = ""
    cur_num = None
    for ch in s:
        if ch in CHOICE_NUM:
            if cur_num is not None:
                parts.append((cur_num, cur.strip()))
            cur_num = CHOICE_NUM.index(ch) + 1
            cur = ""
        else:
            cur += ch
    if cur_num is not None:
        parts.append((cur_num, cur.strip()))
    # Sort by num and return values
    parts.sort()
    # Pad to 5
    out = ["", "", "", "", ""]
    for num, val in parts:
        if 1 <= num <= 5:
            out[num - 1] = val
    return out


def sql_str(s):
    if s is None:
        return "NULL"
    return "'" + str(s).replace("\\", "\\\\").replace("'", "''") + "'"


def sql_jsonb(obj):
    if obj is None:
        return "NULL"
    return "'" + json.dumps(obj, ensure_ascii=False).replace("'", "''") + "'::jsonb"


def parse_intent(text_block):
    """'문제 의도 ...\\n한국어[\\n베트남어]\\n\\n보기 해설...'
    원본은 두 가지 포맷이 섞여 있음:
      - 포맷 A: '문제 의도 Ý định của câu hỏi\\n한국어\\n베트남어' (소수)
      - 포맷 B: '문제 의도\\n한국어' (다수, 베트남어 없음)
    → (intent_ko, intent_vi, choice_explanations dict)"""
    if not text_block:
        return None, None, None
    s = str(text_block)
    intent_ko = intent_vi = None
    choice_explanations = {}

    # Split into 의도 / 해설 sections
    intent_section = ""
    explanation_section = ""
    if "보기 해설" in s:
        parts = s.split("보기 해설", 1)
        intent_section = parts[0]
        explanation_section = parts[1] if len(parts) > 1 else ""
    else:
        intent_section = s

    # 의도 — "문제 의도" 헤더 (한·베 둘 다, 또는 한국어만) 제거 후 본문 추출
    body = intent_section
    if "Ý định của câu hỏi" in body:
        body = body.split("Ý định của câu hỏi", 1)[1]
    elif "문제 의도" in body:
        body = body.split("문제 의도", 1)[1]

    # body 의 paragraph 단위 (빈 줄 기준) — 첫 단락이 의도
    paragraphs = re.split(r"\n\s*\n", body.strip())
    if paragraphs and paragraphs[0].strip():
        first_para = paragraphs[0].strip()
        # 단락 안의 줄들 — 한·베 분리
        para_lines = [l.strip() for l in first_para.split("\n") if l.strip()]
        if len(para_lines) == 1:
            # 한 줄만 있으면 한국어
            intent_ko = para_lines[0]
        elif len(para_lines) >= 2:
            # 두 줄 이상 — 한국어 / 베트남어 추정
            # 한글 (가-힣) 포함 여부로 판단
            first_has_ko = bool(re.search(r"[가-힣]", para_lines[0]))
            second_has_ko = bool(re.search(r"[가-힣]", para_lines[1]))
            if first_has_ko and not second_has_ko:
                intent_ko = para_lines[0]
                intent_vi = para_lines[1]
            elif not first_has_ko and second_has_ko:
                intent_vi = para_lines[0]
                intent_ko = para_lines[1]
            else:
                # 둘 다 한글이면 합치기 (또는 첫 줄)
                intent_ko = "\n".join(para_lines)

    # 보기 해설: "(1) ko\nvi\n(2) ko\nvi\n..."
    if explanation_section:
        # Find each (N) ... block
        blocks = re.split(r"\((\d+)\)", explanation_section)
        # blocks[0] is preamble (likely "Giải thích lựa chọn"), then alternating num/content
        for i in range(1, len(blocks) - 1, 2):
            num = blocks[i].strip()
            content = blocks[i + 1].strip() if i + 1 < len(blocks) else ""
            choice_explanations[num] = content

    return intent_ko, intent_vi, choice_explanations or None


def parse_key_terms(text_block):
    """'용어1: 한국어 정의\nVi: vietnamese def\n\n용어2: ...' → list of dicts"""
    if not text_block:
        return None
    s = str(text_block).strip()
    if not s:
        return None
    # Split by double newlines
    blocks = re.split(r"\n\s*\n", s)
    out = []
    for block in blocks:
        lines = [l.strip() for l in block.strip().split("\n") if l.strip()]
        if not lines:
            continue
        # First line: "용어: 한국어 정의"
        # Second line: "Term: Vietnamese definition"
        ko_term = ko_def = vi_term = vi_def = None
        if lines[0].count(":") >= 1:
            parts = lines[0].split(":", 1)
            ko_term = parts[0].strip()
            ko_def = parts[1].strip() if len(parts) > 1 else None
        if len(lines) >= 2 and lines[1].count(":") >= 1:
            parts = lines[1].split(":", 1)
            vi_term = parts[0].strip()
            vi_def = parts[1].strip() if len(parts) > 1 else None
        out.append(
            {
                "term_ko": ko_term,
                "term_vi": vi_term,
                "def_ko": ko_def,
                "def_vi": vi_def,
            }
        )
    return out or None


def main():
    if len(sys.argv) != 3:
        print(__doc__, file=sys.stderr)
        sys.exit(1)

    questions_path = Path(sys.argv[1])
    analysis_path = Path(sys.argv[2])

    # 분석 자료 먼저 dict 로
    wb_a = openpyxl.load_workbook(analysis_path, data_only=True)
    ws_a = wb_a.active
    analysis_by_id = {}
    for row in ws_a.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        qid = int(row[0])
        analysis_by_id[qid] = {
            "explanation_block": row[2],
            "key_terms_block": row[4],
        }

    # 문제 리스트
    wb_q = openpyxl.load_workbook(questions_path, data_only=True)
    ws_q = wb_q.active

    print("-- =============================================================================")
    print(
        "-- 0012_cbt_questions_data.sql (auto-generated by scripts/import_cbt.py)"
    )
    print("-- =============================================================================")
    print()
    print("delete from public.cbt_questions;")
    print()

    inserted = 0
    skipped = 0
    for row in ws_q.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        qid = int(row[0])
        question = row[1]
        choices_text = row[2]
        answer = row[3]
        chapter_raw = row[4]

        if not question or answer is None:
            skipped += 1
            continue

        chapter = normalize_chapter(chapter_raw)
        choices = split_choices(choices_text)
        if not choices or all(c == "" for c in choices):
            skipped += 1
            continue

        try:
            answer_idx = int(float(answer))
        except (ValueError, TypeError):
            skipped += 1
            continue

        analysis = analysis_by_id.get(qid)
        intent_ko = intent_vi = None
        choice_explanations = None
        key_terms = None
        if analysis:
            intent_ko, intent_vi, choice_explanations = parse_intent(
                analysis["explanation_block"]
            )
            key_terms = parse_key_terms(analysis["key_terms_block"])

        print(
            "insert into public.cbt_questions "
            "(id, chapter, question, choices, answer_index, "
            "intent_ko, intent_vi, choice_explanations, key_terms) "
            f"values ({qid}, {sql_str(chapter)}, {sql_str(question)}, "
            f"{sql_jsonb(choices)}, {answer_idx}, "
            f"{sql_str(intent_ko)}, {sql_str(intent_vi)}, "
            f"{sql_jsonb(choice_explanations)}, {sql_jsonb(key_terms)});"
        )
        inserted += 1

    print()
    print(f"-- inserted: {inserted}")
    print(f"-- skipped: {skipped}")


if __name__ == "__main__":
    main()
